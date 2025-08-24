#!/usr/bin/env python3
"""
Test template (default) workflow



Please note that you'll need to replace `'your_file.xlsx'` with the actual path to your Excel file. Also, the `openpyxl` library is used for Excel file manipulation in Python. You can install it using `pip install openpyxl` if you haven't already.
"""
import openpyxl
from openpyxl.comments import Comment

def get_context_info(context_ref):
    # This function should be implemented based on the actual logic of GetContextInfo in the original code
    # For demonstration purposes, it returns a placeholder string
    return f"Context info for {context_ref}"

def get_dimensions_info(context_ref):
    # This function should be implemented based on the actual logic of GetDimensionsInfo in the original code
    # For demonstration purposes, it returns a placeholder string
    return f"Dimensions info for {context_ref}"

def generate_audit_information(file_path):
    """
    Generate audit information as Excel comments.
    
    Args:
    file_path (str): The path to the Excel file.
    """
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Set the instance document to zero
    started_at_sheet = wb.active.title
    
    # Iterate over the rows in the "Mapping" sheet
    mapping_sheet = wb["Mapping"]
    empty_cell_count = 0
    i_counter = 0
    
    for row in range(2, mapping_sheet.max_row + 1):
        i_counter += 1
        cell_value = mapping_sheet.cell(row=row, column=1).value
        
        if cell_value is None:
            # Cell is empty
            empty_cell_count += 1
        else:
            s_spreadsheet = cell_value
            s_cell = mapping_sheet.cell(row=row, column=2).value
            s_concept = mapping_sheet.cell(row=row, column=3).value
            s_context_ref = mapping_sheet.cell(row=row, column=4).value
            s_unit_ref = mapping_sheet.cell(row=row, column=5).value
            s_decimals = mapping_sheet.cell(row=row, column=6).value
            s_scale = mapping_sheet.cell(row=row, column=7).value
            
            # Get the context info
            txt_context_info = get_context_info(s_context_ref)
            txt_dimensions_info = get_dimensions_info(s_context_ref)
            
            # Get the value from the specified sheet and cell
            sheet = wb[s_spreadsheet]
            cell = sheet[s_cell]
            s_value = cell.value
            
            # Create a comment
            comment = Comment(f"~~~~~~~~~~~~~~~~~~XBRL Fact Value Details~~~~~~~~~~~~~~~~~~\n"
                              f"Concept: {s_concept}\n"
                              f"Context ID: {s_context_ref}\n", 
                              "Author")
            
            if s_unit_ref:
                # Is numeric
                comment.text += (f"Units: {s_unit_ref}\n"
                                 f"Decimals: {s_decimals}\n"
                                 f"Fact Value: {s_value}\n"
                                 f"Scale: {s_scale}\n\n"
                                 f"XBRL: \n"
                                 f"<{s_concept} contextRef='{s_context_ref}' unitRef='{s_unit_ref}' decimals='{s_decimals}'>{s_value}</{s_concept}>\n\n"
                                 f"~~~~~~~~~~~~~~~~~~Context Details~~~~~~~~~~~~~~~~~~~~~\n"
                                 f"{txt_context_info}\n"
                                 f"{txt_dimensions_info}\n\n")
            else:
                # Is text
                comment.text += (f"Fact Value: {s_value}\n\n"
                                 f"XBRL: \n"
                                 f"<{s_concept} contextRef='{s_context_ref}'>{s_value}</{s_concept}>\n\n"
                                 f"~~~~~~~~~~~~~~~~~~Context Details~~~~~~~~~~~~~~~~~~~~~\n"
                                 f"{txt_context_info}\n"
                                 f"{txt_dimensions_info}\n\n")
            
            # Add the comment to the cell
            cell.comment = comment
            
        if empty_cell_count == 100:
            break
    
    # Save the workbook
    wb.save(file_path)

# Example usage
generate_audit_information("example.xlsx")