#!/usr/bin/env python3
"""
Test template (default) workflow



This Python code uses the `openpyxl` library to interact with Excel workbooks, which is the most popular Python library for working with Excel files. The code follows the same logical flow as the original VBA code, but with Python's syntax and conventions.

"""

__author__ = "Jason Beach"
__version__ = "0.1.0"
__license__ = "AGPL-3.0"


import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

def generate_xbrl():
    # Get the active workbook and sheet
    wb = openpyxl.load_workbook(filename='your_workbook.xlsx')  # Replace with your workbook path
    started_at_sheet = wb.active.title

    # Disable screen updating (equivalent to Application.ScreenUpdating = False in VBA)
    # Note: This is not directly translatable in Python as it's an Excel-specific feature

    # Select the "Mapping" worksheet
    ws_mapping = wb['Mapping']

    # Get the output file path
    output_file = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(output_file, wb['Setup']['B2'].value)

    if not output_file:
        # Cancel button pressed
        return

    # Create the output file
    with open(output_file, 'w', encoding='utf-8') as ts:
        # Root element
        ts.write('<?xml version="1.0" encoding="utf-8"?>\n')
        ts.write('<!-- HelloWorld Example -->\n')
        ts.write(f'<!-- Date/time created: {datetime.now()} -->\n')

        # Writes the initial group with all the namespace declarations and schemaLocations
        ts.write("<xbrl  xmlns='http://www.xbrl.org/2003/instance'\n")
        ts.write("       xmlns:xbrli = 'http://www.xbrl.org/2003/instance'\n")
        ts.write("       xmlns:link = 'http://www.xbrl.org/2003/linkbase'\n")
        ts.write("       xmlns:xlink = 'http://www.w3.org/1999/xlink'\n")
        ts.write("       xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'\n")
        ts.write("       xmlns:iso4217='http://www.xbrl.org/2003/iso4217'\n")

        # Go grab the schemas....
        ws_taxonomies = wb['Taxonomies']

        # Write namespace declarations
        for row in ws_taxonomies.iter_rows(min_row=2, max_col=2):
            if row[0].value and len(str(row[0].value)) > 1:
                ts.write(f"       xmlns:{row[0].value}='{row[1].value}'\n")

        ts.write("       xsi:schemaLocation='\n")

        # Write schema locations
        for row in ws_taxonomies.iter_rows(min_row=2, max_col=5):
            if row[0].value and len(str(row[0].value)) > 1 and row[4].value == "Yes":
                ts.write(f"                 {row[1].value} {row[2].value} \n")

        ts.write("       '>\n")
        ts.write("\n")

        # Write schemaRefs
        for row in ws_taxonomies.iter_rows(min_row=2, max_col=5):
            if row[0].value and len(str(row[0].value)) > 1 and row[3].value == "Yes":
                ts.write(f"   <link:schemaRef xlink:type='simple' xlink:href='{row[2].value}' />\n")

        # Process contexts....
        ws_contexts = wb['Contexts']

        ts.write("\n")
        ts.write("   <!-- Contexts -->\n")

        for row in ws_contexts.iter_rows(min_row=2, max_col=6):
            if row[0].value and len(str(row[0].value)) > 1:
                ts.write(f"   <context id='{row[0].value}'>\n")
                ts.write("      <entity>\n")
                ts.write(f"         <identifier scheme='{row[2].value}'>{row[1].value}</identifier>\n")
                ts.write("      </entity>\n")
                ts.write("      <period>\n")

                if row[3].value == "Instant":
                    ts.write(f"         <instant>{row[5].value}</instant>\n")
                elif row[3].value == "Duration":
                    ts.write(f"         <startDate>{row[4].value}</startDate>\n")
                    ts.write(f"         <endDate>{row[5].value}</endDate>\n")
                else:
                    ts.write("         <instant>ERROR</instant>\n")

                ts.write("      </period>\n")
                ts.write("   </context>\n")

        # Process units
        ws_units = wb['Units']

        ts.write("\n")
        ts.write("   <!-- Units -->\n")

        for row in ws_units.iter_rows(min_row=2, max_col=2):
            if row[0].value and len(str(row[0].value)) > 1:
                ts.write(f"   <unit id='{row[0].value}'>\n")
                ts.write(f"      <measure>{row[1].value}</measure>\n")
                ts.write("   </unit>\n")

        # Fact values output based on mapping information
        ws_mapping = wb['Mapping']

        empty_cell_count = 0

        ts.write("\n")
        ts.write("   <!-- Fact values -->\n")

        for row in ws_mapping.iter_rows(min_row=2):
            if not row[0].value:
                empty_cell_count += 1
                if empty_cell_count == 100:
                    break
            else:
                empty_cell_count = 0
                # Process the row data here
                # This part of the code was incomplete in the original VBA code
                # You'll need to implement the specific logic for processing the mapping data

if __name__ == "__main__":
    generate_xbrl()
