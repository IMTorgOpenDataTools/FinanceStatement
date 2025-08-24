#!/usr/bin/env python3
"""
Test template (default) workflow




"""
import openpyxl
from openpyxl import Workbook

# Initialize message number
m_MessageNumber = 0

def validate_hello_world(file_path):
    """
    Validates the Schedule worksheet in the given Excel file.
    
    Args:
        file_path (str): Path to the Excel file.
    """
    global m_MessageNumber
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the worksheet
    ws = wb['Hello World Example']
    
    # 2007 PPE Totals
    m_MessageNumber += 1
    lngPPE = int(ws['C14'].value)
    lngPPESum = int(ws['C9'].value) + int(ws['C10'].value) + int(ws['C11'].value) + int(ws['C12'].value) + int(ws['C13'].value)
    
    # Create a new worksheet for validation results
    validation_ws = wb.create_sheet('Validation Results')
    
    # Write the validation result for 2007 PPE
    if lngPPE == lngPPESum:
        # PPE is OK!
        validation_ws.cell(row=m_MessageNumber, column=1).value = m_MessageNumber
        validation_ws.cell(row=m_MessageNumber, column=2).value = "Information"
        validation_ws.cell(row=m_MessageNumber, column=3).value = "Property, Plant and Equipment for 2007 foots on 'Hello World Example'."
    else:
        # Error in 2007 PPE
        validation_ws.cell(row=m_MessageNumber, column=1).value = m_MessageNumber
        validation_ws.cell(row=m_MessageNumber, column=2).value = "ERROR"
        validation_ws.cell(row=m_MessageNumber, column=3).value = f"Property, Plant and Equipment for 2007 does not foot!  Total Property, Plant and Equipment for group = {lngPPE}, whereas the sum of the components of PPE is {lngPPESum}. Please correct discrepency on 'Hello World Example'."
    
    # 2006 PPE Totals
    m_MessageNumber += 1
    lngPPE = int(ws['D14'].value)
    lngPPESum = int(ws['D9'].value) + int(ws['D10'].value) + int(ws['D11'].value) + int(ws['D12'].value) + int(ws['D13'].value)
    
    # Write the validation result for 2006 PPE
    if lngPPE == lngPPESum:
        # PPE is OK!
        validation_ws.cell(row=m_MessageNumber, column=1).value = m_MessageNumber
        validation_ws.cell(row=m_MessageNumber, column=2).value = "Information"
        validation_ws.cell(row=m_MessageNumber, column=3).value = "Property, Plant and Equipment for 2006 foots on 'Hello World Example'."
    else:
        # Error in 2006 PPE
        validation_ws.cell(row=m_MessageNumber, column=1).value = m_MessageNumber
        validation_ws.cell(row=m_MessageNumber, column=2).value = "ERROR"
        validation_ws.cell(row=m_MessageNumber, column=3).value = f"Property, Plant and Equipment for 2006 does not foot!  Total Property, Plant and Equipment for group = {lngPPE}, whereas the sum of the components of PPE is {lngPPESum}. Please correct discrepency on 'Hello World Example'."
    
    # Save the workbook
    wb.save(file_path)

# Example usage
validate_hello_world('example.xlsx')