#!/usr/bin/env python3
"""
Entrypoint for creating instance document


Note: This code assumes that the Excel file is in the same directory as the Python script. 
You'll need to replace `"example.xlsx"` with the actual path to your Excel file and `"example_context_ref"` 
with the actual context reference you're searching for. Also, make sure to install the `openpyxl` '
'library if you haven't already, you can do this by running `pip install openpyxl` in your terminal.
"""

__author__ = "Jason Beach"
__version__ = "0.1.0"
__license__ = "AGPL-3.0"





import openpyxl
from openpyxl import load_workbook

# Load the workbook
def load_workbook_file(file_path):
    """Load the workbook from the given file path."""
    return load_workbook(filename=file_path)

# Get context info
def get_context_info(worksheet, s_context_ref):
    """
    Get context info from the worksheet based on the given context reference.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to search in.
        s_context_ref (str): The context reference to search for.

    Returns:
        str: The context info as a string.
    """
    context_info = ""
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0] == s_context_ref:
            context_info += f"Context ID: {row[0]}\n"
            context_info += f"Entity Scheme: {row[2]}\n"
            context_info += f"Entity ID: {row[1]}\n"
            if row[3] == "Instant":
                context_info += f"Period: [As of] {row[5]}\n"
            elif row[3] == "Duration":
                context_info += f"Period: [For Period] {row[4]} to {row[5]}\n"
            context_info += f"Period: {row[3]}\n"
            context_info += "\n"
            break
    return context_info

# Get dimensions info
def get_dimensions_info(worksheet, s_context_ref):
    """
    Get dimensions info from the worksheet based on the given context reference.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to search in.
        s_context_ref (str): The context reference to search for.

    Returns:
        str: The dimensions info as a string.
    """
    dimensions_info = ""
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0] == s_context_ref:
            dimensions_info += "\n"
            dimensions_info += f"Dimension: {row[2]}\n"
            dimensions_info += f"Member: {row[3]}\n"
    return dimensions_info

# Example usage
if __name__ == "__main__":
    file_path = "example.xlsx"  # Replace with your file path
    wb = load_workbook_file(file_path)
    worksheet = wb.active
    s_context_ref = "example_context_ref"  # Replace with your context reference

    context_info = get_context_info(worksheet, s_context_ref)
    dimensions_info = get_dimensions_info(worksheet, s_context_ref)

    print(context_info)
    print(dimensions_info)