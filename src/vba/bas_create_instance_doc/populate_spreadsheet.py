#!/usr/bin/env python3
"""
Test template (default) workflow



### Notes:
1. Replace `'your_workbook.xlsx'` with your actual Excel file path.
2. The code uses `openpyxl` for Excel operations and `lxml` for XML parsing.
3. Error handling is simplified compared to the VB version but covers the main cases.
4. The XML parsing is more robust with proper error handling.
5. The tuple-related variables are included but not fully implemented as their usage wasn't clear in the original code.
6. The code maintains the same logic flow as the original VB code while following Python best practices.
"""
import os
from lxml import etree
from openpyxl import load_workbook

def populate_spreadsheet():
    # Load the workbook
    wb = load_workbook(filename=os.path.abspath('your_workbook.xlsx'), data_only=True)
    ws_setup = wb["Setup"]
    ws_mapping = wb["Mapping"]

    # Get the starting sheet name
    started_at_sheet = wb.active.title

    # Determine file path
    if str(ws_setup['B5'].value).startswith("http://"):
        file_path = ws_setup['B5'].value
    else:
        file_path = os.path.join(os.path.dirname(os.path.abspath(wb.filename)), ws_setup['B2'].value)

    print(f"Loading data from: {file_path}")

    try:
        # Parse XML document
        parser = etree.XMLParser(recover=True)
        o_document = etree.parse(file_path, parser)

        # Check for XML errors
        if o_document.getroot() is None:
            raise ValueError("Failed to parse XML document")

        empty_cell_count = 0
        tuple_status = False
        tuple_element = ""
        tuple_number = 0
        i_counter = 0

        # Start from row 2 in Mapping sheet
        current_row = 2

        while empty_cell_count < 100:
            i_counter += 1

            # Check if current row is empty
            if ws_mapping[f'A{current_row}'].value is None:
                empty_cell_count += 1
            else:
                empty_cell_count = 0  # Reset counter if we find a non-empty cell

                s_value = "[error]"
                s_spreadsheet = ws_mapping[f'A{current_row}'].value
                s_cell = ws_mapping[f'B{current_row}'].value
                s_concept = ws_mapping[f'C{current_row}'].value
                s_context_ref = ws_mapping[f'D{current_row}'].value
                s_unit_ref = ws_mapping[f'E{current_row}'].value
                s_decimals = ws_mapping[f'F{current_row}'].value
                s_scale = ws_mapping[f'G{current_row}'].value

                # Handle numeric vs text values
                if s_unit_ref and s_decimals is not None:
                    # Numeric value
                    node = o_document.find(f".//{s_concept}[@{'contextRef'}='{s_context_ref}']")
                    if node is not None and node.text is not None:
                        try:
                            s_value = float(node.text) / float(s_scale)
                        except (ValueError, TypeError):
                            s_value = "[error]"
                else:
                    # Text value
                    node = o_document.find(f".//{s_concept}[@{'contextRef'}='{s_context_ref}']")
                    if node is not None:
                        s_value = node.text if node.text is not None else "[error]"

                # Write to target sheet
                if s_spreadsheet in wb.sheetnames:
                    target_sheet = wb[s_spreadsheet]
                    target_sheet[s_cell] = s_value

            current_row += 1

    except Exception as e:
        print(f"Error loading file: {file_path}\nError: {str(e)}")
        return

    finally:
        # Save and close workbook
        wb.save(filename=os.path.abspath('your_workbook.xlsx'))
        wb.close()

        # Return to original sheet (conceptual - Python doesn't have active sheet state)
        print(f"Processing complete. Started at sheet: {started_at_sheet}")

if __name__ == "__main__":
    populate_spreadsheet()