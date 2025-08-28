#!/usr/bin/env python3
"""
Test template (default) workflow


Note: This code assumes that you have the `openpyxl` library installed. You can install it using pip: `pip install openpyxl`. Also, replace `"input.xlsx"` with the path to your Excel file. The output will be saved as `"output.xlsx"` in the same directory. The `validate_hello_world` function is left empty and should be implemented according to your needs.
"""
import openpyxl
from openpyxl import Workbook

class Validator:
    def __init__(self, file_path):
        # Load the workbook
        self.wb = openpyxl.load_workbook(file_path)
        # Select the "Messages" worksheet
        self.ws = self.wb["Messages"]
        # Initialize the message number
        self.m_MessageNumber = 1

    def run_validation(self):
        # Delete all existing messages
        self.ws.delete_rows(2, 500)
        
        # Set the first message
        self.ws.cell(row=2, column=1).value = self.m_MessageNumber
        self.ws.cell(row=2, column=2).value = "Information"
        self.ws.cell(row=2, column=3).value = "Validation started"

        # Call the validation function
        self.validate_hello_world()

        # Set the completion message
        self.m_MessageNumber += 1
        self.ws.cell(row=3, column=1).value = self.m_MessageNumber
        self.ws.cell(row=3, column=2).value = "Information"
        self.ws.cell(row=3, column=3).value = "Validation complete"

        # Save the workbook
        self.wb.save("output.xlsx")

    def validate_hello_world(self):
        # This function should be implemented according to your needs
        pass

# Example usage
if __name__ == "__main__":
    validator = Validator("input.xlsx")
    validator.run_validation()